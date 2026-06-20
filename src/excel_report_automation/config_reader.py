"""macro_manager.xlsx の macro_setting シートを読み込むモジュール。"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

SHEET_NAME = "macro_settings"


@dataclass
class MacroTask:
    order: int
    enabled: bool
    file_path: Path
    macro_name: str
    timeout_min: int
    on_error: str  # "continue" or "stop"
    notes: str


def load_tasks(manager_path: str | Path) -> list[MacroTask]:
    """macro_manager.xlsx から有効なタスクを順番に読み込む。"""
    manager_path = Path(manager_path).resolve()
    if not manager_path.exists():
        raise FileNotFoundError(f"設定ファイルが見つかりません: {manager_path}")

    wb = openpyxl.load_workbook(manager_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"シート '{SHEET_NAME}' が見つかりません: {manager_path}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    # 1行目=ヘッダー、2行目=日本語ラベル、3行目以降=データ
    data_rows = rows[2:]

    tasks = []
    for row in data_rows:
        if not row or row[0] is None:
            continue

        try:
            order = int(row[0])
        except (ValueError, TypeError):
            continue

        _, enabled, file_path, macro_name, timeout_min, on_error, notes = row

        tasks.append(MacroTask(
            order=order,
            enabled=(enabled == 1),
            file_path=Path(str(file_path)),
            macro_name=str(macro_name),
            timeout_min=int(timeout_min) if timeout_min else 5,
            on_error=str(on_error).lower() if on_error else "stop",
            notes=str(notes) if notes else "",
        ))

    tasks.sort(key=lambda t: t.order)
    logger.info("全タスク数: %d", len(tasks))
    return tasks
